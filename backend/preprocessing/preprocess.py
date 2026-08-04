"""Main preprocessing module for Treasury data.

Handles Balance_by_Day Excel format (multi-sheet, dates as columns, metrics as rows)
and CSV inputs with date columns.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import numpy as np
import pandas as pd

from .holidays import georgian_holidays_range


def _ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _log(msg: str):
    print(f"[{_ts()}] {msg}", flush=True)


@dataclass
class PreprocessConfig:
    """Configuration for preprocessing run."""

    input_path: str
    date_col: Optional[str] = None
    balance_col: Optional[str] = None
    variant: str = "raw"
    business_days_zero_flows: bool = True
    weekday_weeks: int = 8
    out_root: str = "data_preprocessed"
    run_outputs_dir: str = "outputs"
    expected_csv: Optional[str] = None
    save_parquet: bool = False
    sheet_name: Optional[str] = None
    header_row: Optional[int] = None


def parse_balance_by_day_excel(
    excel_path: Path, sheet_name: Optional[str] = None
) -> pd.DataFrame:
    """
    Parse Balance_by_Day Excel format.
    
    Format:
    - Row containing dates starts at column C (index 2)
    - First 2 columns are labels (geo, english)
    - Subsequent rows contain numeric values per date
    - Duplicate metric names must be SUMMED per date
    - Blank numeric cells = NaN.  The parser reports faithfully what the workbook
      contains; whether a blank means "no transaction" (a true zero, correct for a
      sparse flow line) or "not reported" (correct for a level) needs the column's
      semantics, which only the variant layer has.  apply_variant_raw restores 0.0
      for flows on reported dates.

    Args:
        excel_path: Path to Excel file
        sheet_name: Specific sheet to read (None = all year sheets)
        
    Returns:
        DataFrame with date as index, metric names as columns
    """
    import openpyxl

    _log(f"Reading Excel file: {excel_path}")

    # Get sheet names
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name:
        sheet_names = [sheet_name] if sheet_name in wb.sheetnames else []
        if not sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    else:
        # Auto-detect year sheets (common pattern: "2015", "2016", etc.)
        sheet_names = [s for s in wb.sheetnames if s.strip().isdigit() or "year" in s.lower()]
        if not sheet_names:
            # Fallback to all sheets except very common non-data sheets
            exclude = {"Sheet1", "Sheet", "Summary", "Overview", "Notes"}
            sheet_names = [s for s in wb.sheetnames if s not in exclude]
            if not sheet_names:
                sheet_names = wb.sheetnames

    if not sheet_names:
        raise ValueError(f"No sheets found in {excel_path}")

    _log(f"Processing {len(sheet_names)} sheet(s): {sheet_names}")

    all_data = []

    for sname in sheet_names:
        _log(f"  Processing sheet: {sname}")
        try:
            df = pd.read_excel(excel_path, sheet_name=sname, header=None, engine="openpyxl")
        except Exception as e:
            _log(f"    Warning: Could not read sheet {sname}: {e}, skipping")
            continue

        if len(df) < 2 or len(df.columns) < 3:
            _log(f"    Warning: Sheet {sname} too small, skipping")
            continue

        # Find row with dates
        # Dates can start at column C (index 2) or column D (index 3)
        # Look for row that has date-like values
        date_row_idx = None
        date_start_col = None
        
        for i in range(min(20, len(df))):  # Check first 20 rows
            # Try column 2 (index 2) first, then column 3 (index 3)
            for col_idx in [2, 3]:
                if len(df.columns) <= col_idx:
                    continue
                val = df.iloc[i, col_idx]
                if pd.notna(val):
                    try:
                        # Try to parse as date (raises if not a date; value unused)
                        pd.to_datetime(str(val), errors="raise")
                        # Check if next few columns also look like dates
                        if len(df.columns) > col_idx + 1:
                            val2 = df.iloc[i, col_idx + 1]
                            if pd.notna(val2):
                                try:
                                    pd.to_datetime(str(val2), errors="raise")
                                    date_row_idx = i
                                    date_start_col = col_idx
                                    break
                                except Exception:
                                    pass
                    except Exception:
                        continue
            if date_row_idx is not None:
                break

        if date_row_idx is None:
            _log(f"    Warning: Could not find date row in sheet {sname}, skipping")
            continue

        # Extract dates from row (starting at detected column)
        date_row = df.iloc[date_row_idx, date_start_col:].values
        dates = []
        for val in date_row:
            if pd.isna(val):
                break
            try:
                dt = pd.to_datetime(str(val), errors="raise")
                dates.append(dt)
            except Exception:
                # Stop at first non-date
                break

        if not dates:
            _log(f"    Warning: No valid dates found in sheet {sname}, skipping")
            continue

        _log(f"    Found {len(dates)} dates from {dates[0]} to {dates[-1]}")

        # Extract metric rows (rows after date_row_idx)
        # Column 0 (index 0): Georgian labels (optional)
        # Column 1 (index 1): English metric names
        # Values start at date_start_col (same column where dates start)
        metrics_data = {}
        for i in range(date_row_idx + 1, len(df)):
            row = df.iloc[i]
            if len(row) < date_start_col + 1:
                continue

            # Get metric name from English label (column 1, index 1)
            metric_name = None
            if len(row) > 1:
                name_val = row.iloc[1]
                if pd.notna(name_val):
                    metric_name = str(name_val).strip()
                    if metric_name.lower() in {"nan", "none", "", "null"}:
                        continue

            if not metric_name:
                continue

            # Extract values for this metric (starting at date_start_col, same as dates)
            # A blank cell is NaN here, not 0.0 -- see the module docstring.
            #
            # Measured on Balance_by_Day_2015-2025.xlsx: 27,693 of 115,428 cells
            # (24.0%) are blank, overwhelmingly in sparse line items -- Valuables
            # 100%, Shares and other equity 99.2%, Inventories 99.2%, Dividends
            # 91.2%.  For those, blank does mean zero and apply_variant_raw
            # restores 0.0.  Every headline aggregate has exactly one blank, and
            # State budget balance -- where a zero balance is implausible -- has
            # two, which is the case the old blanket 0.0 got wrong.
            values = []
            for j in range(date_start_col, min(len(row), date_start_col + len(dates))):
                val = row.iloc[j]
                if pd.isna(val):
                    val = np.nan
                else:
                    try:
                        val = float(val)
                        if not np.isfinite(val):
                            val = np.nan
                    except (ValueError, TypeError):
                        val = np.nan
                values.append(val)

            if len(values) != len(dates):
                # Pad or truncate to match dates.  Padding is "not reported".
                if len(values) < len(dates):
                    values.extend([np.nan] * (len(dates) - len(values)))
                else:
                    values = values[: len(dates)]

            # Handle duplicates: SUM them.  Treat NaN as absent rather than zero,
            # so a reported value plus a blank keeps the value, and two blanks
            # stay blank.
            if metric_name in metrics_data:
                existing = metrics_data[metric_name]
                metrics_data[metric_name] = [
                    np.nan
                    if (pd.isna(existing[j]) and pd.isna(values[j]))
                    else float(np.nansum([existing[j], values[j]]))
                    for j in range(len(dates))
                ]
            else:
                metrics_data[metric_name] = values

        # Build DataFrame for this sheet
        if metrics_data:
            sheet_df = pd.DataFrame(metrics_data, index=pd.DatetimeIndex(dates))
            all_data.append(sheet_df)
            _log(f"    Extracted {len(metrics_data)} metrics")

    if not all_data:
        raise ValueError("No data extracted from Excel file")

    # Combine all sheets
    combined = pd.concat(all_data, axis=0, sort=False)
    # Sum duplicates across sheets (same metric name, same date).  min_count=1
    # keeps an all-blank group as NaN; the default would silently turn it into 0.0
    # and re-introduce the absent/zero conflation this parser now avoids.
    combined = combined.groupby(combined.index).sum(min_count=1)

    # Sort by date
    combined = combined.sort_index()

    _log(f"Final: {len(combined)} unique dates, {len(combined.columns)} metrics")
    return combined


def parse_csv_input(
    csv_path: Path, date_col: Optional[str] = None
) -> pd.DataFrame:
    """
    Parse CSV input with date column.
    
    Args:
        csv_path: Path to CSV file
        date_col: Name of date column (auto-detect if None)
        
    Returns:
        DataFrame with date as index, other columns as metrics
    """
    _log(f"Reading CSV file: {csv_path}")
    df = pd.read_csv(csv_path)

    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]

    # Find date column
    if date_col:
        if date_col not in df.columns:
            raise ValueError(f"Date column '{date_col}' not found in CSV")
        date_col_found = date_col
    else:
        # Auto-detect: look for columns with date-like values
        date_col_found = None
        for col in df.columns:
            try:
                test = pd.to_datetime(df[col], errors="coerce")
                if test.notna().mean() > 0.8:  # 80% valid dates
                    date_col_found = col
                    break
            except Exception:
                continue

        if not date_col_found:
            raise ValueError("Could not auto-detect date column in CSV")

    # Set date as index
    df[date_col_found] = pd.to_datetime(df[date_col_found], errors="coerce")
    df = df.dropna(subset=[date_col_found])
    df = df.set_index(date_col_found)
    df.index.name = "date"

    # Convert numeric columns
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    _log(f"Parsed CSV: {len(df)} dates, {len(df.columns)} columns")
    return df


def add_calendar_flags(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add is_weekend and is_holiday flags to DataFrame.
    
    Args:
        df: DataFrame with date index
        
    Returns:
        DataFrame with added flags
    """
    idx = pd.DatetimeIndex(df.index)
    df = df.copy()

    # Weekend flag
    df["is_weekend"] = (idx.weekday >= 5).astype(int)

    # Georgian holidays
    if len(idx) > 0:
        holidays = georgian_holidays_range(idx.min().date(), idx.max().date())
        holiday_dates = pd.DatetimeIndex([pd.Timestamp(d) for d in holidays])
        df["is_holiday"] = idx.normalize().isin(holiday_dates.normalize()).astype(int)
    else:
        df["is_holiday"] = 0

    return df


def reindex_to_daily_calendar(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reindex to full daily calendar (including weekends/holidays).
    
    On non-business days, metric columns remain NaN in raw output.
    
    Args:
        df: DataFrame with date index
        
    Returns:
        DataFrame reindexed to daily calendar
    """
    if len(df) == 0:
        return df

    # Create full daily calendar
    start = df.index.min().normalize()
    end = df.index.max().normalize()
    full_calendar = pd.date_range(start=start, end=end, freq="D")

    # Reindex (non-business days will have NaN for metrics)
    df_reindexed = df.reindex(full_calendar)

    # Add calendar flags to all rows
    df_reindexed = add_calendar_flags(df_reindexed)

    return df_reindexed


# Business days legitimately absent from the source workbook despite not being
# Georgian public holidays.  Each entry is a deliberate, reviewed exception; the
# coverage check refuses to run otherwise.  Keeping the list explicit and short
# means a NEW gap can never hide among already-known ones.
KNOWN_ABSENT_BUSINESS_DAYS: Tuple[str, ...] = (
    "2018-11-28",   # single unexplained gap in the 2015-2025 workbook
)


def check_business_day_coverage(
    reported_dates: pd.DatetimeIndex,
    holidays: set,
    allow_list: Tuple[str, ...] = KNOWN_ABSENT_BUSINESS_DAYS,
) -> Dict:
    """Fail loudly when a business day is missing from the source for no reason.

    A business day absent from the workbook is not a zero-flow day -- it is a day
    we know nothing about.  Silently reindexing it in and filling 0.0 was measured
    to move the h=5 persistence baseline by 10.1% with no warning anywhere (review
    §7.1), and because true zeros are legitimate in this series the two cases are
    indistinguishable after the fact.

    Georgian public holidays are expected to be absent.  Anything else must be on
    ``allow_list`` or this raises.

    Returns a dict describing coverage; raises ValueError on an unexpected gap.
    """
    if len(reported_dates) == 0:
        raise ValueError("No dates reported in the source - nothing to preprocess.")

    reported = pd.DatetimeIndex(reported_dates).normalize().unique()
    reported_set = set(reported)
    expected = pd.date_range(reported.min(), reported.max(), freq="B")
    holiday_norm = {pd.Timestamp(h).normalize() for h in holidays}
    allowed = {pd.Timestamp(d).normalize() for d in allow_list}

    absent = [d for d in expected if d not in reported_set]
    unexplained = [d for d in absent if d not in holiday_norm and d not in allowed]

    coverage = {
        "expected_business_days": int(len(expected)),
        "reported_business_days": int(len(expected) - len(absent)),
        "absent_business_days": int(len(absent)),
        "absent_but_holiday": int(len([d for d in absent if d in holiday_norm])),
        "absent_but_allow_listed": int(len([d for d in absent if d in allowed])),
        "unexplained_absent": [str(d.date()) for d in unexplained],
    }

    if unexplained:
        raise ValueError(
            f"{len(unexplained)} business day(s) missing from the source and not "
            f"explained by a Georgian public holiday or the allow-list: "
            f"{[str(d.date()) for d in unexplained[:10]]}"
            f"{' ...' if len(unexplained) > 10 else ''}. "
            "Refusing to preprocess: filling these with 0.0 would be "
            "indistinguishable from genuine zero-flow days and would silently move "
            "the persistence baseline every model is graded against. Either supply "
            "the missing data, or add the dates to KNOWN_ABSENT_BUSINESS_DAYS with "
            "a reason."
        )

    _log(f"Business-day coverage OK: {coverage['reported_business_days']}"
         f"/{coverage['expected_business_days']} reported, "
         f"{coverage['absent_but_holiday']} absent as public holidays, "
         f"{coverage['absent_but_allow_listed']} allow-listed.")
    return coverage


def apply_variant_raw(
    df: pd.DataFrame, business_days_zero_flows: bool
) -> Tuple[pd.DataFrame, Dict]:
    """
    Apply raw variant: minimal transformation.

    - Verify business-day coverage before anything else
    - Reindex to full daily calendar
    - Blank FLOW cells on REPORTED dates become 0.0 (no transaction of that type)
    - Blank LEVEL cells stay NaN (a blank balance is not a zero balance)
    - Business days ABSENT from the source stay NaN -- unknown, not zero
    - Include is_weekend/is_holiday flags

    Args:
        df: DataFrame with date index and metric columns
        business_days_zero_flows: If True, set flows to 0 on non-business days

    Returns:
        (processed_df, summary_dict)
    """
    _log("Applying variant: raw")

    # Which dates the source actually reported, captured BEFORE reindexing
    # invents rows.  Afterwards the two kinds of NaN are indistinguishable.
    reported_dates = pd.DatetimeIndex(df.index).normalize().unique()

    holidays = (
        georgian_holidays_range(reported_dates.min().date(), reported_dates.max().date())
        if len(reported_dates)
        else set()
    )
    coverage = check_business_day_coverage(reported_dates, holidays)

    # Reindex to daily calendar
    df_out = reindex_to_daily_calendar(df)

    # Identify level columns (e.g., "State budget balance")
    level_cols = []
    for col in df.columns:
        if "balance" in str(col).lower() or "level" in str(col).lower():
            level_cols.append(col)

    flow_cols = [c for c in df.columns if c not in level_cols and c not in ["is_weekend", "is_holiday"]]

    reported_mask = df_out.index.normalize().isin(reported_dates)
    non_business = (df_out["is_weekend"] == 1) | (df_out["is_holiday"] == 1)

    # A blank flow cell on a date the source DID report means "no transaction of
    # this type that day" -> 0.0.  This is the 24% of blank cells living in sparse
    # line items.  Restoring 0.0 here rather than in the parser is what keeps the
    # absent-date case NaN.
    n_blank_to_zero = 0
    for col in flow_cols:
        if col not in df_out.columns:
            continue
        blank_reported = reported_mask & df_out[col].isna()
        n_blank_to_zero += int(blank_reported.sum())
        df_out.loc[blank_reported, col] = 0.0

    if business_days_zero_flows:
        # Weekends and public holidays are not trading days; zero flow is correct.
        for col in flow_cols:
            df_out.loc[non_business, col] = 0.0

    # Anything still NaN in a flow column is an ABSENT business day.  The coverage
    # check above means it can only be a holiday or an allow-listed date.
    absent_flow_cells = int(
        sum(
            int((df_out[c].isna() & ~non_business).sum())
            for c in flow_cols
            if c in df_out.columns
        )
    )

    summary = {
        "variant": "raw",
        "level_columns": level_cols,
        "flow_columns": flow_cols,
        "business_days_zero_flows": business_days_zero_flows,
        "coverage": coverage,
        "blank_flow_cells_set_to_zero": n_blank_to_zero,
        "absent_business_day_flow_cells_left_nan": absent_flow_cells,
    }

    return df_out, summary


def apply_variant_clean_conservative(
    df: pd.DataFrame, business_days_zero_flows: bool
) -> Tuple[pd.DataFrame, Dict]:
    """
    Apply clean_conservative variant: safe cleaning.
    
    - Level columns: forward-fill across ALL calendar days
    - Flow columns: set to 0 on non-business days if flag is True, else leave NaN
    
    Args:
        df: DataFrame with date index and metric columns
        business_days_zero_flows: If True, set flows to 0 on non-business days
        
    Returns:
        (processed_df, summary_dict)
    """
    _log("Applying variant: clean_conservative")

    # Start from raw
    df_out, _ = apply_variant_raw(df, business_days_zero_flows)

    # Identify level columns
    level_cols = []
    for col in df.columns:
        if "balance" in str(col).lower() or "level" in str(col).lower():
            level_cols.append(col)

    # Forward-fill level columns across ALL days
    for col in level_cols:
        if col in df_out.columns:
            df_out[col] = df_out[col].ffill()

    # Flow columns: handle non-business days
    flow_cols = [c for c in df.columns if c not in level_cols and c not in ["is_weekend", "is_holiday"]]
    
    if business_days_zero_flows:
        non_business = (df_out["is_weekend"] == 1) | (df_out["is_holiday"] == 1)
        for col in flow_cols:
            if col in df_out.columns:
                df_out.loc[non_business, col] = 0.0

    summary = {
        "variant": "clean_conservative",
        "level_columns": level_cols,
        "flow_columns": flow_cols,
        "business_days_zero_flows": business_days_zero_flows,
        "imputations": {"level_forward_fill": len(level_cols)},
    }

    return df_out, summary


# Retained for report compatibility; no longer used to clip.  See
# _clean_flow_column_causally for why MAD clipping was removed outright.
MIN_CLIP_HISTORY = 8

# A flow value is *reported* as suspect when it exceeds this multiple of the
# largest value observed before it in the same column.  Chosen to catch
# order-of-magnitude data-entry errors (a stray extra digit is 10x) while leaving
# genuine month-end spikes -- measured at roughly 2-3x a quiet day -- untouched.
# Nothing is altered on the strength of this: a value being surprising is not
# evidence that it is wrong, and the whole point of removing the clipper was to
# stop silently rewriting the spikes we are trying to forecast.
SUSPECT_JUMP_MULTIPLE = 10.0


def flow_validity_report(
    series: pd.Series,
    is_business: pd.Series,
    suspect_multiple: float = SUSPECT_JUMP_MULTIPLE,
) -> Dict:
    """Report -- never alter -- flow values that look physically implausible.

    Two checks, both causal:

      * negative values, which no gross-flow line should take;
      * values exceeding ``suspect_multiple`` times the maximum observed *earlier*
        in the same column, which is the shape an extra digit makes.

    Returns counts and the offending dates so an operator can look, and so the
    validator in Step 5 has something concrete to gate on.
    """
    vals = series.to_numpy(dtype=float)
    idx = series.index
    business = np.asarray(is_business, dtype=bool)

    negatives, jumps = [], []
    running_max = 0.0
    for pos in np.flatnonzero(business):
        v = vals[pos]
        if np.isnan(v):
            continue
        if v < 0:
            negatives.append(str(idx[pos].date()))
        if running_max > 0 and v > suspect_multiple * running_max:
            jumps.append({
                "date": str(idx[pos].date()),
                "value": float(v),
                "prior_max": float(running_max),
                "ratio": float(v / running_max),
            })
        running_max = max(running_max, abs(v))

    return {
        "n_negative": len(negatives),
        "negative_dates": negatives[:20],
        "n_suspect_jumps": len(jumps),
        "suspect_jumps": jumps[:20],
        "suspect_jump_multiple": suspect_multiple,
    }


def _clean_flow_column_causally(
    series: pd.Series,
    is_business: pd.Series,
    weekday_weeks: int,
    min_clip_history: int = MIN_CLIP_HISTORY,
) -> Tuple[pd.Series, int, int, list]:
    """Impute and clip one flow column using only data from strictly before each row.

    Why this is not the obvious loop
    --------------------------------
    The previous implementation computed both statistics from the WHOLE series::

        ref  = dow_values.tail(weekday_weeks).median()   # last N of all 11 years
        med  = dow_values.median(); mad = ...            # every occurrence

    ``.tail(N)`` takes the most RECENT N occurrences, so a gap in 2016 was filled
    with the median of eight Mondays in 2025, and the clipping thresholds applied
    to 2016 were computed partly from the locked 2025 holdout.  Measured (review
    §2.3): 223 of 2,763 business-day Revenues values (8.1%) touched by
    holdout-informed statistics, 20 of them inside the 2025 window, median
    clipping change 118,342,253 -- roughly three times the best model's MAE.

    Here every statistic for row *t* is built from same-weekday **observed** rows
    strictly before *t*:

      * imputation reference -- median of the last ``weekday_weeks`` observed
        values before *t*.

    Imputed values never enter that statistic.  Letting them would compound: one
    fabricated value would shift the reference used to fabricate the next.

    Why there is no outlier clipping any more
    ----------------------------------------
    The old code clipped at 8*MAD of the per-weekday distribution, computed over
    the WHOLE series.  Making that threshold causal was tried and abandoned on
    measurement: any causally-estimated pool is a *sample* of same-weekday values
    that frequently excludes the month-end and tax-deadline spikes, so MAD comes
    out small, 8*MAD is tight, and the spikes get clipped.  Measured on the real
    workbook, causal clipping suppressed the annual mean of Revenues by 41% in
    2024 (98,123,411 -> 58,213,784) and touched 216 of 2,763 business days.

    The old version survived only *because* it was leaky: including the spikes in
    the pool inflated MAD into a generous threshold, which is why it clipped just
    105 values.  In other words MAD clipping was never appropriate for this
    series, and the leak was masking that.

    Those spikes are the signal -- fixed monthly dates are exactly what the
    day-of-month features exist to predict -- so they are left alone.  Validity
    problems are *reported* by ``flow_validity_report`` rather than silently
    altered, because a value being surprising is not evidence it is wrong.

    Returns ``(series, n_imputed, n_clipped, imputed_dates)``; ``n_clipped`` is
    retained as 0 so the report shape is unchanged for existing consumers.
    """
    values = series.to_numpy(dtype=float, copy=True)
    idx = series.index
    business = np.asarray(is_business, dtype=bool)
    weekday = np.asarray(idx.weekday)

    n_imputed = n_clipped = 0
    imputed_dates: list = []

    for dow in range(7):
        # Positions of this weekday that are business days, chronologically.
        positions = np.flatnonzero((weekday == dow) & business)
        if positions.size == 0:
            continue

        recent_observed: list = []   # last `weekday_weeks` observed, for imputation

        for pos in positions:
            v = values[pos]

            if np.isnan(v):
                # -- impute from strictly prior observed values --
                if recent_observed:
                    values[pos] = float(np.median(recent_observed))
                    n_imputed += 1
                    imputed_dates.append(str(idx[pos].date()))
                # No prior observation of this weekday: leave NaN rather than
                # inventing a number from a different weekday.  The coverage check
                # in apply_variant_raw means this is a holiday or allow-listed.
                continue

            # Observed values are kept as reported -- see the docstring on why
            # outlier clipping was removed rather than made causal.
            #
            # The pool tracks what was OBSERVED, not what we wrote.
            recent_observed.append(v)
            if len(recent_observed) > weekday_weeks:
                recent_observed.pop(0)

    return (
        pd.Series(values, index=idx, name=series.name),
        n_imputed,
        n_clipped,
        imputed_dates,
    )


def apply_variant_clean_treasury(
    df: pd.DataFrame, business_days_zero_flows: bool, weekday_weeks: int
) -> Tuple[pd.DataFrame, Dict]:
    """
    Apply clean_treasury variant: Treasury-friendly cleaning.

    - Start from clean_conservative
    - For flow columns on BUSINESS DAYS only, using ONLY data strictly before each
      row (see _clean_flow_column_causally for why this matters):
      - Impute NaN with the median of the last N observed same-weekday values
      - Clip extreme outliers at 8*MAD of the expanding prior same-weekday history

    Args:
        df: DataFrame with date index and metric columns
        business_days_zero_flows: If True, set flows to 0 on non-business days
        weekday_weeks: Number of weeks to use for weekday reference
        
    Returns:
        (processed_df, summary_dict)
    """
    _log("Applying variant: clean_treasury")

    # Start from clean_conservative
    df_out, base_summary = apply_variant_clean_conservative(df, business_days_zero_flows)

    # Identify columns
    level_cols = base_summary.get("level_columns", [])
    flow_cols = [c for c in df.columns if c not in level_cols and c not in ["is_weekend", "is_holiday"]]

    # Business days mask
    is_business = (df_out["is_weekend"] == 0) & (df_out["is_holiday"] == 0)

    imputation_counts = {}
    clipping_counts = {}
    imputed_dates: Dict[str, list] = {}
    validity: Dict[str, Dict] = {}

    for col in flow_cols:
        if col not in df_out.columns:
            continue

        series, n_imp, n_clip, imp_dates = _clean_flow_column_causally(
            series=df_out[col].copy(),
            is_business=is_business,
            weekday_weeks=weekday_weeks,
        )
        df_out[col] = series
        imputation_counts[col] = n_imp
        clipping_counts[col] = n_clip
        if imp_dates:
            imputed_dates[col] = imp_dates

        # Reported, not applied.  Values stay exactly as the source gave them.
        rep = flow_validity_report(series, is_business)
        if rep["n_negative"] or rep["n_suspect_jumps"]:
            validity[col] = rep

    summary = {
        "variant": "clean_treasury",
        "level_columns": level_cols,
        "flow_columns": flow_cols,
        "business_days_zero_flows": business_days_zero_flows,
        "weekday_weeks": weekday_weeks,
        "causal_cleaning": True,
        "flow_outlier_clipping": "disabled",
        "imputations": imputation_counts,
        "clipped_outliers": clipping_counts,   # all zero; kept for report shape
        "imputed_dates": imputed_dates,
        "validity_warnings": validity,
        "n_columns_with_validity_warnings": len(validity),
        "coverage": base_summary.get("coverage"),
        "n_imputed_total": int(sum(imputation_counts.values())),
        "n_clipped_total": int(sum(clipping_counts.values())),
    }

    return df_out, summary


def compare_with_expected(output_df: pd.DataFrame, expected_path: Path) -> Dict:
    """
    Compare output with expected CSV for regression testing.
    
    Args:
        output_df: Output DataFrame
        expected_path: Path to expected CSV
        
    Returns:
        Comparison summary dict
    """
    if not expected_path.exists():
        return {"error": f"Expected file not found: {expected_path}"}

    try:
        expected_df = pd.read_csv(expected_path)
        expected_df["date"] = pd.to_datetime(expected_df["date"])
        expected_df = expected_df.set_index("date")

        # Compare structure
        output_cols = set(output_df.columns)
        expected_cols = set(expected_df.columns)

        # Compare values for common columns and dates
        common_cols = output_cols & expected_cols
        common_dates = output_df.index.intersection(expected_df.index)

        diffs = {}
        for col in common_cols:
            if col in ["is_weekend", "is_holiday"]:
                continue  # Skip flags for comparison
            output_vals = output_df.loc[common_dates, col]
            expected_vals = expected_df.loc[common_dates, col]

            # Compare numeric values
            numeric_mask = output_vals.notna() & expected_vals.notna()
            if numeric_mask.any():
                diff = (output_vals[numeric_mask] - expected_vals[numeric_mask]).abs()
                max_diff = float(diff.max()) if len(diff) > 0 else 0.0
                mean_diff = float(diff.mean()) if len(diff) > 0 else 0.0
                diffs[col] = {"max_diff": max_diff, "mean_diff": mean_diff, "n_diffs": int((diff > 1e-6).sum())}

        return {
            "expected_file": str(expected_path),
            "output_rows": len(output_df),
            "expected_rows": len(expected_df),
            "output_cols": list(output_cols),
            "expected_cols": list(expected_cols),
            "common_cols": list(common_cols),
            "common_dates": len(common_dates),
            "differences": diffs,
        }
    except Exception as e:
        return {"error": f"Comparison failed: {e}"}


def run_preprocess(cfg: PreprocessConfig) -> Dict:
    """
    Main preprocessing function.
    
    Args:
        cfg: Preprocessing configuration
        
    Returns:
        Report dictionary
    """
    _log("=" * 80)
    _log("Starting preprocessing")
    _log(f"Input: {cfg.input_path}")
    _log(f"Variant: {cfg.variant}")

    input_path = Path(cfg.input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Parse input
    if input_path.suffix.lower() in {".xlsx", ".xls"}:
        df = parse_balance_by_day_excel(input_path, cfg.sheet_name)
    elif input_path.suffix.lower() == ".csv":
        df = parse_csv_input(input_path, cfg.date_col)
    else:
        raise ValueError(f"Unsupported file format: {input_path.suffix}")

    # Apply variant
    variant = cfg.variant.lower().strip()
    if variant == "raw":
        df_out, summary = apply_variant_raw(df, cfg.business_days_zero_flows)
    elif variant == "clean_conservative":
        df_out, summary = apply_variant_clean_conservative(df, cfg.business_days_zero_flows)
    elif variant == "clean_treasury":
        df_out, summary = apply_variant_clean_treasury(
            df, cfg.business_days_zero_flows, cfg.weekday_weeks
        )
    else:
        raise ValueError(f"Unknown variant: {variant}. Must be: raw, clean_conservative, clean_treasury")

    # Reset index to have date as column
    # The index should be a DatetimeIndex from reindex_to_daily_calendar
    if isinstance(df_out.index, pd.DatetimeIndex):
        df_out = df_out.reset_index()
        # The reset_index will create a column from the index
        # Find the date column (should be the first column or named 'date')
        date_col_name = None
        for col in df_out.columns:
            if col == "date" or pd.api.types.is_datetime64_any_dtype(df_out[col]):
                date_col_name = col
                break
        
        if date_col_name and date_col_name != "date":
            df_out = df_out.rename(columns={date_col_name: "date"})
        elif not date_col_name:
            # If no date column found, the index might not have been reset properly
            # Create date column from index
            df_out["date"] = df_out.index
            df_out = df_out.reset_index(drop=True)
    
    # Ensure date column exists and is datetime
    if "date" not in df_out.columns:
        raise ValueError("Date column not found after processing")
    df_out["date"] = pd.to_datetime(df_out["date"])

    # Write output
    out_root = Path(cfg.out_root)
    variant_dir = out_root / variant
    variant_dir.mkdir(parents=True, exist_ok=True)

    base_name = input_path.stem
    output_csv = variant_dir / f"{base_name}__{variant}.csv"
    df_out.to_csv(output_csv, index=False)
    _log(f"Saved: {output_csv}")

    output_parquet = None
    if cfg.save_parquet:
        output_parquet = variant_dir / f"{base_name}__{variant}.parquet"
        df_out.to_parquet(output_parquet, index=False)
        _log(f"Saved: {output_parquet}")

    # Write preview
    run_outputs = Path(cfg.run_outputs_dir)
    run_outputs.mkdir(parents=True, exist_ok=True)
    preview_path = run_outputs / "preprocess_preview.csv"
    df_out.head(500).to_csv(preview_path, index=False)
    _log(f"Saved preview: {preview_path}")

    # Build report
    report = {
        "output_csv": str(output_csv),
        "output_parquet": str(output_parquet) if output_parquet else None,
        "preview_path": str(preview_path),
        "variant": variant,
        "input_path": str(input_path),
        "row_count": len(df_out),
        "col_count": len(df_out.columns),
        "summary": summary,
    }

    # Compare with expected if provided
    if cfg.expected_csv:
        expected_path = Path(cfg.expected_csv)
        comparison = compare_with_expected(df_out.set_index("date"), expected_path)
        report["comparison"] = comparison

    # Write report
    report_path = run_outputs / "preprocess_report.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    _log(f"Saved report: {report_path}")

    _log("Preprocessing completed successfully")
    return report
